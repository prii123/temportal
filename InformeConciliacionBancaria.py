import pandas as pd
# Crear un nuevo archivo Excel y escribir en él usando openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
#  PARA LOS PDFS
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle


class InformeConciliacionBancaria:
    def __init__(
            self,
            nombre_archivo_generado,
            montosBancarios_entradas,
            montosBancarios_salidas,
            montosContables_entradas,
            montosContables_salidas,
            informacionEmpresa,
            informacionConciliada,
            duplicados_entradas_resumen,
            duplicados_salidas_resumen
    ):
        self.nombre_archivo_generado = nombre_archivo_generado
        self.montosBancarios_entradas = montosBancarios_entradas
        self.montosBancarios_salidas = montosBancarios_salidas
        self.montosContables_entradas = montosContables_entradas
        self.montosContables_salidas = montosContables_salidas
        self.informacionEmpresa = informacionEmpresa
        self.informacionConciliada = informacionConciliada
        self.duplicados_entradas_resumen = duplicados_entradas_resumen
        self.duplicados_salidas_resumen = duplicados_salidas_resumen
        #print(informacionConciliada.get('conciliado_entradas'))

        partidasBancarias_entradas = {
            'fecha': montosBancarios_entradas['fecha_x'],
            'Descripcion': montosBancarios_entradas['descripcion_x'],
            'Valor': montosBancarios_entradas['valor_x'],
            'Tipo': montosBancarios_entradas['tipo_x'],
            'ID': montosBancarios_entradas['id']
        }
        partidasBancarias_salidas = {
            'fecha': montosBancarios_salidas['fecha_x'],
            'Descripcion': montosBancarios_salidas['descripcion_x'],
            'Valor': montosBancarios_salidas['valor_x'],
            'Tipo': montosBancarios_salidas['tipo_x'],
            'ID': montosBancarios_salidas['id']
        }
        partidasContables_entradas = {
            'fecha': montosContables_entradas['fecha_y'],
            'Descripcion': montosContables_entradas['descripcion_y'],
            'Valor': montosContables_entradas['valor_y'],
            'Tipo': montosContables_entradas['tipo_y'],
            'ID': montosContables_entradas['id']
        }
        partidasContables_salidas = {
            'fecha': montosContables_salidas['fecha_y'],
            'Descripcion': montosContables_salidas['descripcion_y'],
            'Valor': montosContables_salidas['valor_y'],
            'Tipo': montosContables_salidas['tipo_y'],
            'ID': montosContables_salidas['id']
        }

        self.df_Bancarios_entradas = pd.DataFrame(partidasBancarias_entradas)
        self.df_Bancarios_salidas = pd.DataFrame(partidasBancarias_salidas)

        self.df_Contable_entradas = pd.DataFrame(partidasContables_entradas)
        self.df_Contable_salidas = pd.DataFrame(partidasContables_salidas)


        # Crear un libro de trabajo de Excel
        self.wb = Workbook()
        # Eliminar la hoja por defecto
        self.wb.remove_sheet(self.wb.active)
        self.sheet = self.wb.create_sheet('Partidas Conciliatorias')
        # Crear otra nueva hoja de trabajo
        self.sheet_elementos_conciliados = self.wb.create_sheet('Elementos Conciliados')
        self.sheet_duplicados = self.wb.create_sheet('Elementos Duplicados')

        #self.config()

    def configuracion_hoja(self, hoja,A=None,B=None,C=None,D=None,E=None,F=None,G=None,H=None,I=None,J=None,K=None,L=None,M=None,N=None,O=None,P=None):
        # Activar la hoja de trabajo
        #self.sheet.title = 'Archivo'
        hoja.border = None  # Esto quita el borde de la celda
        # Ocultar las líneas de cuadrícula
        hoja.sheet_view.showGridLines = False
        # Define el ancho de cada columna
        hoja.column_dimensions['A'].width = A or 15
        hoja.column_dimensions['B'].width = B or 15
        hoja.column_dimensions['C'].width = C or 15
        hoja.column_dimensions['D'].width = D or 15
        hoja.column_dimensions['E'].width = E or 15
        hoja.column_dimensions['F'].width = F or 15
        hoja.column_dimensions['G'].width = G or 15
        hoja.column_dimensions['H'].width = H or 15
        hoja.column_dimensions['I'].width = I or 15
        hoja.column_dimensions['J'].width = J or 15
        hoja.column_dimensions['K'].width = K or 15
        hoja.column_dimensions['L'].width = L or 15
        hoja.column_dimensions['M'].width = M or 15
        hoja.column_dimensions['N'].width = N or 15
        hoja.column_dimensions['O'].width = O or 15
        hoja.column_dimensions['P'].width = P or 15


    def cabeceras(self, hoja):
        hoja.merge_cells('A1:G1')
        cell = hoja.cell(row=1, column=1, value=self.informacionEmpresa.get('razonSocial'))
        cell.font = Font(bold=True)
        #cell.border = None  # Quitar el borde de la celda
        #cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')

        hoja.merge_cells('A2:G2')
        cell = hoja.cell(row=2, column=1, value=self.informacionEmpresa.get('nit'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)

        hoja.merge_cells('A3:G3')
        cell = hoja.cell(row=3, column=1, value=self.informacionEmpresa.get('periodo'))
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
    def sumarColumna(self, data):
        suma = 0
        try:
            for i in data['Valor']:
                suma += i
            return suma
        except:
            return suma

    def hoja_elementos_conciliados(self):
        # Activar la hoja de trabajo
        self.configuracion_hoja(self.sheet_elementos_conciliados, A=18, B=18, C=32, D=18, E=32, F=18)
        self.cabeceras(self.sheet_elementos_conciliados)

        posicionIteracion = 6
        #print(self.informacionConciliada.get('conciliado_entradas'))
        # MONTOS CONCILIADOS
        self.sheet_elementos_conciliados.merge_cells('A' + str(posicionIteracion) + ':F' + str(posicionIteracion) + '')
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=1, value="MONTOS CONCILIADOS ENTRADAS")
        cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
        cell.alignment=Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        # DIVISION BANCOS Y CONTABLES
        posicionIteracion+=1
        self.sheet_elementos_conciliados.merge_cells('C' + str(posicionIteracion) + ':D' + str(posicionIteracion) + '')
        self.sheet_elementos_conciliados.merge_cells('E' + str(posicionIteracion) + ':F' + str(posicionIteracion) + '')
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=3,
                                                     value="BANCARIOS")
        cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
        cell.alignment=Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=5,
                                                     value="CONTABLE")
        cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
        cell.alignment=Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        #TITULOS
        posicionIteracion += 1
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=1,
                                                     value="ID")
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=2,
                                                     value="FECHA")
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=3,
                                                     value="DESCRIPCION")
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=4,
                                                     value="VALOR")
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=5,
                                                     value="DESCRIPCION")
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=6,
                                                     value="VALOR")
        cell.font = Font(bold=True)


        posicionIteracion += 1
        for row_index, row in self.informacionConciliada.get('conciliado_entradas').iterrows():
            posicionIteracion += 1
            #print(row)
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=1, value=row.get('id')) #id
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=2, value=row.get('fecha_x'))#fecha
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=3, value=row.get('descripcion_x'))#descripcion banco
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=4, value=row.get('valor_x'))#valor banco
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=5, value=row.get('descripcion_y'))#descripcion contable
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=6, value=row.get('valor_y'))#valor contable

        #-----------------------------------------------------------------------------------------------------
        # ITERACION DE LAS SALIDAS
        posicionIteracion += 6
        # MONTOS CONCILIADOS
        self.sheet_elementos_conciliados.merge_cells('A' + str(posicionIteracion) + ':F' + str(posicionIteracion) + '')
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=1, value="MONTOS CONCILIADOS SALIDAS")
        cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
        cell.alignment=Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        # DIVISION BANCOS Y CONTABLES
        posicionIteracion += 1
        self.sheet_elementos_conciliados.merge_cells('C' + str(posicionIteracion) + ':D' + str(posicionIteracion) + '')
        self.sheet_elementos_conciliados.merge_cells('E' + str(posicionIteracion) + ':F' + str(posicionIteracion) + '')
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=3,
                                                     value="BANCARIOS")
        cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
        cell.alignment=Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=5,
                                                     value="CONTABLE")
        cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
        cell.alignment=Alignment(horizontal='center', vertical='center')
        cell.font = Font(bold=True)
        # TITULOS
        posicionIteracion += 1
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=1,
                                                     value="ID")
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=2,
                                                     value="FECHA")
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=3,
                                                     value="DESCRIPCION")
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=4,
                                                     value="VALOR")
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=5,
                                                     value="DESCRIPCION")
        cell.font = Font(bold=True)
        cell = self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=6,
                                                     value="VALOR")
        cell.font = Font(bold=True)

        posicionIteracion += 1
        for row_index, row in self.informacionConciliada.get('conciliado_salidas').iterrows():
            posicionIteracion += 1
            # print(row)
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=1, value=row.get('id'))  # id
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=2, value=row.get('fecha_x'))  # fecha
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=3,
                                                  value=row.get('descripcion_x'))  # descripcion banco
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=4, value=row.get('valor_x'))  # valor banco
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=5,
                                                  value=row.get('descripcion_y'))  # descripcion contable
            self.sheet_elementos_conciliados.cell(row=posicionIteracion, column=6,
                                                  value=row.get('valor_y'))  # valor contable

    def hoja_partidas_conciliatorias(self):
        self.configuracion_hoja(self.sheet, A=18.0, B=40.0, C=20.0, D=15.0, E=15.0)
        self.cabeceras(self.sheet)
        posicionIteracion = 6

        # MONTOS BANCARIOS ITERACIONES
        self.sheet.merge_cells('A'+str(posicionIteracion)+':G'+str(posicionIteracion)+'')
        cell = self.sheet.cell(row=posicionIteracion, column=1, value="Partidas conciliatorias - bancos")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
        posicionIteracion+=2
        cell = self.sheet.cell(row=(posicionIteracion), column=1, value="Entradas")
        cell.font = Font(bold=True)
        cell = self.sheet.cell(row=(posicionIteracion), column=3, value= (self.sumarColumna(self.df_Bancarios_entradas)))
        cell.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
        cell.font = Font(bold=True)

        # ITERACION DE PARTIDAS CONCILIATORIAS ENTRADAS
        #posicionIteracion += 1
        for row_index, row in self.df_Bancarios_entradas.iterrows():
            posicionIteracion += 1
            for col_index, value in enumerate(row, start=1):
                self.sheet.cell(row= posicionIteracion, column=col_index, value=value)
                self.sheet.fill = PatternFill(start_color="CFF3FF", end_color="CFF3FF",
                                        fill_type="solid")

        # ITERACION DE PARTIDAS CONCILIATORIAS SALIDAS
        posicionIteracion += 2
        cell = self.sheet.cell(row = posicionIteracion, column=1, value="Salidas")
        cell.font = Font(bold=True)
        cell = self.sheet.cell(row=(posicionIteracion), column=3,
                               value=(self.sumarColumna(self.df_Bancarios_salidas)))
        cell.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
        cell.font = Font(bold=True)


        for row_index, row in self.df_Bancarios_salidas.iterrows():
            posicionIteracion += 1
            for col_index, value in enumerate(row, start=1):
                 self.sheet.cell(row=posicionIteracion, column=col_index, value=value)
                 self.sheet.fill = PatternFill(start_color="CFF3FF", end_color="CFF3FF",
                                         fill_type="solid")



        #MONTOS CONTABLES ITERACIONES
        posicionIteracion += 3
        self.sheet.merge_cells('A' + str(posicionIteracion) + ':G' + str(posicionIteracion) + '')
        cell = self.sheet.cell(row=(posicionIteracion), column=1, value="Partidas conciliatorias - contable")
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")

        # ITERACION DE PARTIDAS CONCILIATORIAS ENTRADAS CONTABLE
        cell = self.sheet.cell(row=(posicionIteracion + 2), column=1, value="Entradas")
        cell.font = Font(bold=True)
        cell = self.sheet.cell(row=(posicionIteracion + 2), column=3,
                               value=(self.sumarColumna(self.df_Contable_entradas)))
        cell.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
        cell.font = Font(bold=True)

        for row_index, row in self.df_Contable_entradas.iterrows():
            posicionIteracion += 1
            for col_index, value in enumerate(row, start=1):
                self.sheet.cell(row=(posicionIteracion + 2), column=col_index, value=value)
                self.sheet.fill = PatternFill(start_color="CFF3FF", end_color="CFF3FF",
                                              fill_type="solid")

        # ITERACION DE PARTIDAS CONCILIATORIAS SALIDAS CONTABLE
        posicionIteracion += 4
        cell = self.sheet.cell(row= posicionIteracion , column=1, value="Salidas")
        cell.font = Font(bold=True)
        cell = self.sheet.cell(row=(posicionIteracion), column=3,
                               value=(self.sumarColumna(self.df_Contable_salidas)))
        cell.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'
        cell.font = Font(bold=True)

        for row_index, row in self.df_Contable_salidas.iterrows():
            posicionIteracion += 1
            for col_index, value in enumerate(row, start=1):
                self.sheet.cell(row=(posicionIteracion), column=col_index, value=value)
                self.sheet.fill = PatternFill(start_color="CFF3FF", end_color="CFF3FF",
                                              fill_type="solid")


        # FORMATO NUMEROS COLUMNA C
        # Dar formato de contabilidad a una columna específica (en este caso la "C" que se refiere a la tercera columna)
        for row in self.sheet.iter_rows(min_row=2, min_col=3, max_row=self.sheet.max_row, max_col=3):
            for cell in row:
                cell.number_format = '"$"#,##0.00_);[Red]("$"#,##0.00)'

    def hoja_elementos_duplicados(self):
        self.configuracion_hoja(self.sheet_duplicados, A=18, B=20, C=20, D=5, E=5, F=5,G=18,H=18,I=20,J=18,K=5,L=18,M=18,N=20,O=18)
        self.cabeceras(self.sheet_duplicados)
        #print(self.duplicados_salidas_resumen)
        counter=6

        if not self.duplicados_entradas_resumen.empty:
            self.sheet_duplicados['A' + str(counter)] = "ID"
            self.sheet_duplicados['B' + str(counter)] = "Contable_Entradas"
            self.sheet_duplicados['C' + str(counter)] = "Bancos_Entradas"
            counter += 1
            for i in range(len(self.duplicados_entradas_resumen)):
                counter+=1
                self.sheet_duplicados['A' + str(counter)] = self.duplicados_entradas_resumen.iloc[i, 0]
                self.sheet_duplicados['B' + str(counter)] = self.duplicados_entradas_resumen.iloc[i, 1]
                self.sheet_duplicados['C' + str(counter)] = self.duplicados_entradas_resumen.iloc[i, 2]
            #counter+=3
            #print(self.informacionConciliada.get('duplicado_entradas'))
            counter2 = 6
            self.sheet_duplicados.merge_cells('G' + str(counter2) + ':O' + str(counter2))
            cell = self.sheet_duplicados.cell(row=counter2, column=7, value='ENTRADAS')
            cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            counter2 += 1
            for elemento in self.informacionConciliada.get('duplicado_entradas'):
                counter2+=1
                for key, value in elemento.items():
                    #cell=self.sheet_duplicados['F' + str(counter2)] = key
                    self.sheet_duplicados.merge_cells('G' + str(counter2) + ':O' + str(counter2))
                    cell = self.sheet_duplicados.cell(row=counter2, column=7, value=key)
                    cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)
                    counter2 += 1
                    #self.sheet_duplicados['L' + str(counter2)] = 'BANCARIOS'
                    self.sheet_duplicados.merge_cells('L'+ str(counter2)+':O'+ str(counter2))
                    cell = self.sheet_duplicados.cell(row=counter2, column=12, value='BANCARIOS')
                    cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)

                    #self.sheet_duplicados['G' + str(counter2)] = 'CONTABLES'
                    self.sheet_duplicados.merge_cells('G' + str(counter2) + ':J' + str(counter2))
                    cell = self.sheet_duplicados.cell(row=counter2, column=7, value='CONTABLES')
                    cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)

                    counter2+=1
                    if (len(value.get('bancos')) > 0):
                        for row in range(len(value.get('bancos'))):
                            counter2+=1
                            self.sheet_duplicados['L' + str(counter2)] = value.get('bancos').values[row][6]
                            self.sheet_duplicados['M' + str(counter2)] = value.get('bancos').values[row][0]
                            self.sheet_duplicados['N' + str(counter2)] = value.get('bancos').values[row][1]
                            self.sheet_duplicados['O' + str(counter2)] = value.get('bancos').values[row][2]
                    else:
                        self.sheet_duplicados['L' + str(counter2)] = 0


                    if (len(value.get('contable')) > 0):
                        for row in range(len(value.get('contable'))):
                            counter2+=1
                            self.sheet_duplicados['G' + str(counter2)] = value.get('contable').values[row][6]
                            self.sheet_duplicados['H' + str(counter2)] = value.get('contable').values[row][0]
                            self.sheet_duplicados['I' + str(counter2)] = value.get('contable').values[row][1]
                            self.sheet_duplicados['J' + str(counter2)] = value.get('contable').values[row][2]
                    else:
                        self.sheet_duplicados['G' + str(counter2)] = 0

                    #if not df_contables.empty:
                    #    print('-------------------------')
                    #    print(key + ' ************ ', value.get('contable'))


        if not self.duplicados_salidas_resumen.empty:
            counter+=3
            self.sheet_duplicados['A' + str(counter)] = "ID"
            self.sheet_duplicados['B' + str(counter)] = "Frecuencia_Contable_Salidas,"
            self.sheet_duplicados['C' + str(counter)] = "Frecuencia_Bancos_Salidas"

            for i in range(len(self.duplicados_salidas_resumen)):
                counter+=1
                self.sheet_duplicados['A' + str(counter)] = self.duplicados_salidas_resumen.iloc[i, 0]
                self.sheet_duplicados['B' + str(counter)] = self.duplicados_salidas_resumen.iloc[i, 1]
                self.sheet_duplicados['C' + str(counter)] = self.duplicados_salidas_resumen.iloc[i, 2]

            counter2 += 6
            self.sheet_duplicados.merge_cells('G' + str(counter2) + ':O' + str(counter2))
            cell = self.sheet_duplicados.cell(row=counter2, column=7, value='SALIDAS')
            cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(bold=True)
            counter2 += 1
            for elemento in self.informacionConciliada.get('duplicado_salidas'):
                counter2 += 1
                for key, value in elemento.items():
                    # cell=self.sheet_duplicados['F' + str(counter2)] = key
                    self.sheet_duplicados.merge_cells('G' + str(counter2) + ':O' + str(counter2))
                    cell = self.sheet_duplicados.cell(row=counter2, column=7, value=key)
                    cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)
                    counter2 += 1
                    # self.sheet_duplicados['L' + str(counter2)] = 'BANCARIOS'
                    self.sheet_duplicados.merge_cells('L' + str(counter2) + ':O' + str(counter2))
                    cell = self.sheet_duplicados.cell(row=counter2, column=12, value='BANCARIOS')
                    cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)

                    # self.sheet_duplicados['G' + str(counter2)] = 'CONTABLES'
                    self.sheet_duplicados.merge_cells('G' + str(counter2) + ':J' + str(counter2))
                    cell = self.sheet_duplicados.cell(row=counter2, column=7, value='CONTABLES')
                    cell.fill = PatternFill(start_color="C2C2C2", end_color="C2C2C2", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(bold=True)

                    counter2 += 1
                    if (len(value.get('bancos')) > 0):
                        for row in range(len(value.get('bancos'))):
                            counter2 += 1
                            self.sheet_duplicados['L' + str(counter2)] = value.get('bancos').values[row][6]
                            self.sheet_duplicados['M' + str(counter2)] = value.get('bancos').values[row][0]
                            self.sheet_duplicados['N' + str(counter2)] = value.get('bancos').values[row][1]
                            self.sheet_duplicados['O' + str(counter2)] = value.get('bancos').values[row][2]
                    else:
                        self.sheet_duplicados['L' + str(counter2)] = 0

                    if (len(value.get('contable')) > 0):
                        for row in range(len(value.get('contable'))):
                            counter2 += 1
                            self.sheet_duplicados['G' + str(counter2)] = value.get('contable').values[row][6]
                            self.sheet_duplicados['H' + str(counter2)] = value.get('contable').values[row][0]
                            self.sheet_duplicados['I' + str(counter2)] = value.get('contable').values[row][1]
                            self.sheet_duplicados['J' + str(counter2)] = value.get('contable').values[row][2]
                    else:
                        self.sheet_duplicados['G' + str(counter2)] = 0

                    # if not df_contables.empty:
                    #    print('-------------------------')
                    #    print(key + ' ************ ', value.get('contable'))

    def guardar(self):
        # Guardar el archivo Excel
        self.wb.save(self.nombre_archivo_generado+'.xlsx')



"""
class InformeConciliacionBancariaPDF:
    def __init__(self, nombre_archivo, montosBancarios_entradas, montosBancarios_salidas, montosContables_entradas, montosContables_salidas):
        self.nombre_archivo = nombre_archivo
        self.montosBancarios_entradas = montosBancarios_entradas
        self.montosBancarios_salidas = montosBancarios_salidas
        self.montosContables_entradas = montosContables_entradas
        self.montosContables_salidas = montosContables_salidas

    def generar_pdf(self):
        doc = SimpleDocTemplate(self.nombre_archivo, pagesize=letter)
        elements = []

        fechas = []
        descripcion = []
        valores = []

        for i in self.montosBancarios_entradas:
            print(i)
            
            fechas.append(i['fecha_x'])
            descripcion.append(i['descripcion_x'])
            valores.append(i['monto_bancario'])
            

        data = [
            ['Fecha', 'Descripción', 'Valor'],
            *zip(fechas, descripcion, valores),
            # Aquí deberías añadir el resto de tus datos de la misma manera
            
            ["brayan", "vallejos", "10000"],
            ["brayan", "vallejos", "10000"],
            ["brayan", "vallejos", "10000"]
           
        ]

        # Create a table
        table = Table(data)
        style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)])

        table.setStyle(style)
        elements.append(table)

        doc.build(elements)

        return f"Se ha generado el archivo PDF: {self.nombre_archivo}"




a = InformeConciliacionBancaria()

a.cabeceras()
a.procesar()
a.guardar()
"""
